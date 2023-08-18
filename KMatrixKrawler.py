import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import zipfile
import threading
import queue


searching = True

def on_enter_pressed(event):
    threaded_search_term()

def threaded_search_term():
    search_thread = threading.Thread(target=search_term)
    search_thread.start()

def stop_search():
    # Diese Funktion wird aufgerufen, wenn der Abbrechen-Button gedrückt wird
    global searching
    searching = False

def find_in_excel(folder, term, progress_callback):
    all_files = [os.path.join(dirpath, f) for dirpath, _, filenames in os.walk(folder) for f in filenames if f.endswith(('.xlsx', '.xls')) and "Vergleich" not in f and not f.startswith('~')]
    total_files = len(all_files)

    for count, full_path in enumerate(all_files, start=1):
        if not searching:  # Wenn die Suche abgebrochen wurde, brechen Sie die Schleife ab
            break
        rel_path = extract_relevant_path_parts(full_path)
        wrapped_text = wrap_text(f"Analysiere... {rel_path}", 80)
        update_queue.put({"action": "update_label", "text": wrapped_text})

        app.update_idletasks()  # Aktualisiert das GUI, um den aktuellen Dateipfad anzuzeigen
        try:
            try:
                workbook = openpyxl.load_workbook(full_path, data_only=True)
            except Exception as e:
                result_text.insert(tk.END, f"Fehler beim Laden der Datei {full_path}: {e}\n")
                app.update_idletasks()
                continue
            for sheet in workbook:
                try:
                    if sheet.max_row > 0:
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.value and term.lower() in str(cell.value).lower():
                                    yield full_path, sheet.title, cell.row, cell.value  # This returns each found match individually
                except Exception as e:
                    print(f"Problem mit Datei: {full_path}")
                    print(e)
        except zipfile.BadZipFile:
            result_text.insert(tk.END, f"Fehler: {full_path} scheint keine gültige Excel-Datei zu sein.\n")
        app.update_idletasks()
        progress_callback(count/total_files*100)  # Update progress


def search_term():
    folder = folder_entry.get()
    term = search_entry.get()

    if not folder or not term:
        messagebox.showinfo("Info", "Bitte Ordner und Suchbegriff eingeben!")
        return

    def update_progress(value):
        update_queue.put({"action": "update_progress", "value": value})


    progress_var.set(0)
    result_text.delete(1.0, tk.END)

    for match in find_in_excel(folder, term, update_progress):
        update_queue.put({"action": "insert_text", "text": f"Datei: {match[0]}\nSheet: {match[1]}\nZeile: {match[2]}\nZellinhalt: {match[3]}\n\n"})
        app.update_idletasks()  # This ensures the GUI is updated after each result is added
    current_file_label.config(text="Analyse abgeschlossen!")

def wrap_text(text, max_length):
    parts = []
    while len(text) > max_length:
        part = text[:max_length]
        last_space = part.rfind(' ')  # sucht das letzte Leerzeichen, um dort umzubrechen
        if last_space != -1:
            parts.append(text[:last_space])
            text = text[last_space:].lstrip()  # Das .lstrip() entfernt das führende Leerzeichen
        else:
            parts.append(part)
            text = text[max_length:]
    parts.append(text)
    return '\n'.join(parts)

def extract_relevant_path_parts(full_path):
    # Holt den Dateinamen
    file_name = os.path.basename(full_path)

    # Holt den übergeordneten Ordner
    parent_folder = os.path.basename(os.path.dirname(full_path))

    return f"{parent_folder}/{file_name}"


def update_gui_from_queue():
    try:
        while True:
            update = update_queue.get_nowait()
            if update["action"] == "update_label":
                current_file_label.config(text=update["text"])
            elif update["action"] == "insert_text":
                result_text.insert(tk.END, update["text"])
            elif update["action"] == "update_progress":
                progress_var.set(update["value"])
                progress_label.config(text=f"{int(update['value'])} %")
    except queue.Empty:
        pass
    app.after(100, update_gui_from_queue)


update_queue = queue.Queue()
app = tk.Tk()
app.title("K-Matrix Krawler")

frame2 = tk.Frame(app)
frame2.grid(row=0, column=0, padx=10, pady=10, sticky=tk.W+tk.E)

frame = tk.Frame(app)
frame.grid(row=1, column=0, padx=10, pady=10, sticky=tk.W+tk.E)

current_file_label = tk.Label(frame2, text="", anchor="w", height=4)
current_file_label.pack(fill=tk.X)

default_folder = "S:/EE_Elektrik_Elektronik/Vernetzungsdaten/V000_Verbundrelease/E3 1.2_P/Aktuell/K-Matrix"

tk.Label(frame, text="Ordner:").grid(row=0, column=0, sticky=tk.W)
folder_entry = tk.Entry(frame, width=50)
folder_entry.insert(0, default_folder)  # Setzt den Standardordnerpfad
folder_entry.grid(row=0, column=1, pady=(0,10))
tk.Button(frame, text="Ordner auswählen", command=lambda: folder_entry.insert(0, filedialog.askdirectory())).grid(row=0, column=2)

tk.Label(frame, text="Suchbegriff:").grid(row=1, column=0, sticky=tk.W)
search_entry = tk.Entry(frame, width=50)
search_entry.grid(row=1, column=1, pady=(0,10))
search_entry.bind('<Return>', on_enter_pressed)
tk.Button(frame, text="Suchen", command=threaded_search_term).grid(row=1, column=2)

progress_frame = tk.Frame(app)
progress_frame.grid(row=2, column=0, padx=10, pady=5, sticky=tk.W+tk.E)

progress_var = tk.DoubleVar()  # Holds progress value
progress_bar = ttk.Progressbar(progress_frame, variable=progress_var, maximum=100)
progress_bar.grid(row=0, column=0, sticky=tk.W+tk.E)

# Label to display percentage
progress_label = tk.Label(progress_frame, text="0 %")
progress_label.grid(row=0, column=1, padx=(5, 0))  # Added padding to separate from progress bar


# Position result_text correctly
result_text = tk.Text(app, height=15, width=70)
scrollbar = tk.Scrollbar(app, command=result_text.yview)
result_text.config(yscrollcommand=scrollbar.set)
result_text.grid(row=3, column=0, padx=10, pady=5, sticky=tk.W+tk.E+tk.N+tk.S)
scrollbar.grid(row=3, column=1, pady=5, sticky=tk.N+tk.S)

# Button in GUI hinzufügen
cancel_button = tk.Button(app, text="Suche abbrechen", command=stop_search)
cancel_button.grid(row=4, column=0, pady=10)

# Configuring the column weights
# app.columnconfigure(0, weight=1)  # This makes sure the column in main window expands
app.grid_rowconfigure(3, weight=1)
app.grid_columnconfigure(0, weight=1)


progress_frame.columnconfigure(0, weight=1)  # This ensures the progress bar column expands


update_gui_from_queue()
app.mainloop()
